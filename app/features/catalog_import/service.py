"""Generate and import controlled Excel workbooks for business catalogs."""

from __future__ import annotations

import sqlite3
from dataclasses import dataclass
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MAX_FILE_BYTES = 5 * 1024 * 1024
MAX_DATA_ROWS = 5_000
MAX_REPORTED_ERRORS = 10

CATALOG_SCHEMAS: dict[str, tuple[tuple[str, str], ...]] = {
    "clients": (
        ("name", "Nome/Razão Social"),
        ("document", "CNPJ/CPF"),
        ("contact", "Contato"),
        ("email", "E-mail"),
        ("phone", "Telefone"),
        ("address", "Endereço"),
        ("city", "Cidade"),
        ("state", "UF"),
        ("segment", "Segmento"),
        ("notes", "Observações"),
    ),
    "suppliers": (
        ("name", "Nome"),
        ("document", "CNPJ"),
        ("contact", "Contato"),
        ("email", "E-mail R.O"),
        ("phone", "Telefone"),
        ("payment_terms", "Condição comercial"),
        ("notes", "Observações"),
    ),
}


class SpreadsheetImportError(ValueError):
    """Raised when a workbook cannot be safely imported."""


@dataclass(frozen=True)
class ImportResult:
    created: int
    updated: int
    ignored: int
    errors: tuple[str, ...]

    def as_dict(self) -> dict[str, Any]:
        return {
            "created": self.created,
            "updated": self.updated,
            "ignored": self.ignored,
            "errors": list(self.errors),
        }


def schema_for(table: str) -> tuple[tuple[str, str], ...]:
    try:
        return CATALOG_SCHEMAS[table]
    except KeyError as exc:
        raise SpreadsheetImportError(f"Catálogo não permitido: {table}") from exc


def build_template(table: str) -> bytes:
    schema = schema_for(table)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Importação"
    sheet.append([label for _, label in schema])
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:{get_column_letter(len(schema))}1"

    header_fill = PatternFill("solid", fgColor="0B7274")
    for index, cell in enumerate(sheet[1], start=1):
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.column_dimensions[get_column_letter(index)].width = max(18, len(str(cell.value)) + 4)
    sheet.row_dimensions[1].height = 24
    instructions = workbook.create_sheet("Instruções")
    instructions["A1"] = "Como importar"
    instructions["A1"].font = Font(size=16, bold=True, color="081426")
    instructions["A2"] = "1. Preencha a aba Importação sem alterar os cabeçalhos."
    instructions["A3"] = "2. Nome e documento são obrigatórios; mantenha CNPJ/CPF como texto."
    instructions["A4"] = "3. Um documento existente atualiza o cadastro; um novo documento cria o registro."
    instructions["A5"] = "4. Limite de 5.000 linhas e arquivo máximo de 5 MiB."
    instructions.column_dimensions["A"].width = 105

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _cell_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_workbook(table: str, content: bytes) -> list[dict[str, Any]]:
    schema = schema_for(table)
    if not content:
        raise SpreadsheetImportError("A planilha está vazia.")
    if len(content) > MAX_FILE_BYTES:
        raise SpreadsheetImportError("A planilha excede o limite de 5 MiB.")

    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise SpreadsheetImportError("Arquivo Excel inválido ou corrompido.") from exc

    try:
        if "Importação" not in workbook.sheetnames:
            raise SpreadsheetImportError("A aba Importação não foi encontrada.")
        sheet = workbook["Importação"]
        first_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers = [_cell_text(value) for value in first_row]
        expected_labels = [label for _, label in schema]
        if len(headers) != len(expected_labels) or set(headers) != set(expected_labels):
            raise SpreadsheetImportError(
                "Cabeçalhos inválidos. Baixe e utilize o modelo oficial."
            )

        label_to_key = {label: key for key, label in schema}
        rows: list[dict[str, Any]] = []
        for row_number, values in enumerate(
            sheet.iter_rows(min_row=2, values_only=True),
            start=2,
        ):
            normalized_values = [_cell_text(value) for value in values[: len(headers)]]
            normalized_values.extend([""] * (len(headers) - len(normalized_values)))
            if not any(normalized_values):
                continue
            if len(rows) >= MAX_DATA_ROWS:
                raise SpreadsheetImportError("A planilha excede o limite de 5.000 linhas.")
            row = {
                label_to_key[label]: normalized_values[index]
                for index, label in enumerate(headers)
            }
            row["_row_number"] = row_number
            rows.append(row)
        return rows
    finally:
        workbook.close()


def _document_key(value: Any) -> str:
    return "".join(character for character in _cell_text(value) if character.isalnum()).casefold()


def import_rows(
    connection: sqlite3.Connection,
    table: str,
    rows: list[dict[str, Any]],
) -> ImportResult:
    schema = schema_for(table)
    fields = [key for key, _ in schema]
    existing_rows = connection.execute(f"SELECT id, document FROM {table}").fetchall()
    existing_by_document = {
        _document_key(row["document"]): row["id"]
        for row in existing_rows
        if _document_key(row["document"])
    }

    created = 0
    updated = 0
    ignored = 0
    errors: list[str] = []
    seen_documents: set[str] = set()

    with connection:
        for row in rows:
            row_number = int(row.get("_row_number") or 0)
            name = _cell_text(row.get("name"))
            document = _cell_text(row.get("document"))
            document_key = _document_key(document)
            if not name or not document_key:
                ignored += 1
                if len(errors) < MAX_REPORTED_ERRORS:
                    errors.append(f"Linha {row_number}: nome e documento são obrigatórios.")
                continue
            if document_key in seen_documents:
                ignored += 1
                if len(errors) < MAX_REPORTED_ERRORS:
                    errors.append(f"Linha {row_number}: documento repetido na planilha.")
                continue
            seen_documents.add(document_key)

            values = [_cell_text(row.get(field)) for field in fields]
            record_id = existing_by_document.get(document_key)
            if record_id:
                assignments = ",".join(f"{field}=?" for field in fields)
                connection.execute(
                    f"UPDATE {table} SET {assignments} WHERE id=?",
                    (*values, record_id),
                )
                updated += 1
            else:
                columns = ",".join(fields)
                placeholders = ",".join("?" for _ in fields)
                cursor = connection.execute(
                    f"INSERT INTO {table}({columns}) VALUES({placeholders})",
                    values,
                )
                existing_by_document[document_key] = cursor.lastrowid
                created += 1

    return ImportResult(
        created=created,
        updated=updated,
        ignored=ignored,
        errors=tuple(errors),
    )

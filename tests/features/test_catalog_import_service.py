import sqlite3
from io import BytesIO

import pytest
from openpyxl import Workbook, load_workbook

from app.features.catalog_import.service import (
    SpreadsheetImportError,
    build_template,
    import_rows,
    parse_workbook,
)


def workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Importação"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.parametrize(
    ("table", "expected_headers"),
    [
        (
            "clients",
            [
                "Nome/Razão Social",
                "CNPJ/CPF",
                "Contato",
                "E-mail",
                "Telefone",
                "Endereço",
                "Cidade",
                "UF",
                "Segmento",
                "Observações",
            ],
        ),
        (
            "suppliers",
            [
                "Nome",
                "CNPJ",
                "Contato",
                "E-mail R.O",
                "Telefone",
                "Condição comercial",
                "Observações",
            ],
        ),
    ],
)
def test_template_contains_import_and_instruction_sheets(
    table: str,
    expected_headers: list[str],
) -> None:
    workbook = load_workbook(BytesIO(build_template(table)))

    assert workbook.sheetnames == ["Importação", "Instruções"]
    assert [cell.value for cell in workbook["Importação"][1]] == expected_headers
    assert "documento" in str(workbook["Instruções"]["A4"].value).lower()


def test_parse_workbook_maps_labels_and_ignores_empty_rows() -> None:
    content = workbook_bytes(
        [
            "CNPJ/CPF",
            "Nome/Razão Social",
            "E-mail",
            "Cidade",
            "UF",
            "Contato",
            "Telefone",
            "Endereço",
            "Segmento",
            "Observações",
        ],
        [
            ["12.345.678/0001-90", "Cliente Excel", "excel@example.invalid", "São Paulo", "SP", "Ana", "11999990000", "Rua A", "Educação", "Linha válida"],
            [None, None, None, None, None, None, None, None, None, None],
        ],
    )

    rows = parse_workbook("clients", content)

    assert rows == [
        {
            "document": "12.345.678/0001-90",
            "name": "Cliente Excel",
            "email": "excel@example.invalid",
            "city": "São Paulo",
            "state": "SP",
            "contact": "Ana",
            "phone": "11999990000",
            "address": "Rua A",
            "segment": "Educação",
            "notes": "Linha válida",
            "_row_number": 2,
        }
    ]


def test_parse_workbook_rejects_structural_header_error() -> None:
    content = workbook_bytes(["Nome/Razão Social", "CNPJ errado"], [["Cliente", "1"]])

    with pytest.raises(SpreadsheetImportError, match="Cabeçalhos inválidos"):
        parse_workbook("clients", content)


def test_import_rows_creates_updates_and_reports_duplicates() -> None:
    connection = sqlite3.connect(":memory:")
    connection.row_factory = sqlite3.Row
    connection.execute(
        """
        CREATE TABLE clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT, document TEXT, contact TEXT, email TEXT, phone TEXT,
            address TEXT, city TEXT, state TEXT, segment TEXT, notes TEXT
        )
        """
    )
    connection.execute(
        "INSERT INTO clients(name,document,email) VALUES(?,?,?)",
        ("Cliente antigo", "12.345.678/0001-90", "old@example.invalid"),
    )

    result = import_rows(
        connection,
        "clients",
        [
            {
                "name": "Cliente atualizado",
                "document": "12345678000190",
                "email": "new@example.invalid",
                "_row_number": 2,
            },
            {
                "name": "Cliente novo",
                "document": "98.765.432/0001-10",
                "email": "created@example.invalid",
                "_row_number": 3,
            },
            {
                "name": "Duplicado na planilha",
                "document": "98.765.432/0001-10",
                "_row_number": 4,
            },
            {"name": "Sem documento", "document": "", "_row_number": 5},
        ],
    )

    saved = connection.execute(
        "SELECT name,document,email FROM clients ORDER BY id"
    ).fetchall()
    assert [tuple(row) for row in saved] == [
        ("Cliente atualizado", "12345678000190", "new@example.invalid"),
        ("Cliente novo", "98.765.432/0001-10", "created@example.invalid"),
    ]
    assert result.created == 1
    assert result.updated == 1
    assert result.ignored == 2
    assert len(result.errors) == 2


def test_unknown_catalog_is_rejected() -> None:
    with pytest.raises(SpreadsheetImportError, match="não permitido"):
        build_template("products")

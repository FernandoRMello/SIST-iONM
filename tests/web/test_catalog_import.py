import sqlite3
from io import BytesIO

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from tests.conftest import LegacyTestState


def workbook_bytes(headers: list[str], rows: list[list[object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Importação"
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


@pytest.mark.parametrize(
    ("table", "filename", "first_header"),
    [
        ("clients", "modelo_clientes.xlsx", "Nome/Razão Social"),
        ("suppliers", "modelo_fornecedores.xlsx", "Nome"),
    ],
)
def test_admin_downloads_official_excel_models(
    admin_client: TestClient,
    table: str,
    filename: str,
    first_header: str,
) -> None:
    response = admin_client.get(f"/cadastros/{table}/import-template")

    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    assert filename in response.headers["content-disposition"]
    workbook = load_workbook(BytesIO(response.content), read_only=True)
    assert workbook["Importação"]["A1"].value == first_header


def test_import_controls_appear_only_for_clients_and_suppliers(
    admin_client: TestClient,
) -> None:
    clients = admin_client.get("/cadastros/clients")
    suppliers = admin_client.get("/cadastros/suppliers")
    products = admin_client.get("/cadastros/products")

    for response in (clients, suppliers):
        assert "Baixar modelo Excel" in response.text
        assert "Importar planilha" in response.text
        assert 'accept=".xlsx"' in response.text
    assert "Baixar modelo Excel" not in products.text
    assert "Importar planilha" not in products.text


def test_client_import_updates_existing_and_creates_new_record(
    admin_client: TestClient,
    legacy_test_state: LegacyTestState,
) -> None:
    content = workbook_bytes(
        [
            "Nome/Razão Social",
            "CNPJ/CPF",
            "Contato",
            "E-mail",
            "Telefone",
            "Endereço",
            "Cidade",
            "UF",
            "Segmento",
            "Observações",
        ],
        [
            ["Cliente QA Atualizado", "98765432000110", "Compras", "updated@example.invalid", "1130000000", "Rua Atualizada", "São Paulo", "SP", "Educação", "Atualizado via Excel"],
            ["Cliente Novo Excel", "11.222.333/0001-44", "Financeiro", "new@example.invalid", "1140000000", "Rua Nova", "Campinas", "SP", "Corporativo", "Criado via Excel"],
        ],
    )

    response = admin_client.post(
        "/cadastros/clients/import",
        files={
            "file": (
                "clientes.xlsx",
                content,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
        follow_redirects=False,
    )

    assert response.status_code == 303
    assert response.headers["location"] == "/cadastros/clients"
    with sqlite3.connect(legacy_test_state.database_path) as connection:
        updated = connection.execute(
            "SELECT name,email,notes FROM clients WHERE id=?",
            (legacy_test_state.ids["client_id"],),
        ).fetchone()
        created = connection.execute(
            "SELECT name,email FROM clients WHERE document=?",
            ("11.222.333/0001-44",),
        ).fetchone()
    assert updated == (
        "Cliente QA Atualizado",
        "updated@example.invalid",
        "Atualizado via Excel",
    )
    assert created == ("Cliente Novo Excel", "new@example.invalid")

    feedback = admin_client.get("/cadastros/clients")
    assert "1 criado" in feedback.text
    assert "1 atualizado" in feedback.text


def test_supplier_import_requires_admin(
    authenticated_client: TestClient,
) -> None:
    template = authenticated_client.get("/cadastros/suppliers/import-template")
    upload = authenticated_client.post(
        "/cadastros/suppliers/import",
        files={"file": ("fornecedores.xlsx", b"not-a-workbook")},
        follow_redirects=False,
    )

    assert template.status_code == 403
    assert upload.status_code == 403


def test_invalid_extension_reports_error_without_writing(
    admin_client: TestClient,
    legacy_test_state: LegacyTestState,
) -> None:
    with sqlite3.connect(legacy_test_state.database_path) as connection:
        before = connection.execute("SELECT COUNT(*) FROM clients").fetchone()[0]

    response = admin_client.post(
        "/cadastros/clients/import",
        files={"file": ("clientes.csv", b"name,document")},
        follow_redirects=False,
    )

    assert response.status_code == 303
    with sqlite3.connect(legacy_test_state.database_path) as connection:
        after = connection.execute("SELECT COUNT(*) FROM clients").fetchone()[0]
    assert after == before
    feedback = admin_client.get("/cadastros/clients")
    assert "arquivo .xlsx" in feedback.text
